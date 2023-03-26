from email import message
import time
import telebot
from openpyxl import load_workbook
from telebot import types
from telebot.types import InlineKeyboardButton,InlineKeyboardMarkup


bot = telebot.TeleBot("5736642101:AAGoKrn5PjzX1smh9MASToDOcmrbLDNbeRk")

@bot.message_handler(commands=['start'])
def start(message):
    key = InlineKeyboardMarkup()
    key.row_width = 2
    btn1 = InlineKeyboardButton(text=f"- البصرة .",callback_data="basra")
    btn2 = InlineKeyboardButton(text=f"- ناصرية .",callback_data="nasrya")
    btn3 = InlineKeyboardButton(text=f"- بغداد .",callback_data="baghdad")
    btn4 = InlineKeyboardButton(text=f"- واسط .",callback_data="waset")
    btn5 = InlineKeyboardButton(text=f"- كربلاء .",callback_data="karbla")
    key.add(btn1)
    key.add(btn2,btn3)
    key.add(btn4)
    key.add(btn5)
    bot.reply_to(message,f"اختر محافظة من الاسفل ",reply_markup=key)
@bot.callback_query_handler(func=lambda call:True)
def q(call):
    if call.data == "basra":
        mm = bot.send_message(call.message.chat.id,f"- ارسل الاسم الثلاثي مثال:\n احمد كاطع صلعوي .")
        bot.register_next_step_handler(mm,get_basra)
    if call.data == "nasrya":
        mm = bot.send_message(call.message.chat.id,f"- ارسل الاسم الثلاثي مثال:\n احمد كاطع صلعوي .")
        bot.register_next_step_handler(mm,get_nasrya)
    if call.data == "baghdad":
        mm = bot.send_message(call.message.chat.id,f"- ارسل الاسم الثلاثي مثال:\n احمد كاطع صلعوي .")
        bot.register_next_step_handler(mm,get_baghdad)
    if call.data == "waset":
        mm = bot.send_message(call.message.chat.id,f"- ارسل الاسم الثلاثي مثال:\n احمد كاطع صلعوي .")
        bot.register_next_step_handler(mm,get_waset)
        if call.data == "karbla":
        mm = bot.send_message(call.message.chat.id,f"- ارسل الاسم الثلاثي مثال:\n احمد كاطع صلعوي .")
        bot.register_next_step_handler(mm,get_waset)
def get_waset(message):
    bot.send_message(message.chat.id,f"- يتم الان محاوله العثور على العائلة .. قد ياخذ بعض الوقت")

    wb = load_workbook("karbla.xlsx",read_only=True)
    ws = wb.active
    for row in ws.rows:


        name = f"{message.text}"
        if str(name.split()[0]) in str(f"{row[1].value}"):
            ae = f"{row[1].value} {row[2].value} {row[3].value}".replace("","")
            if name == ae:
                bot.reply_to(message,f"تم العثور على الشخص .. يتم الان جلب عائلتة .")
                idF = f'{row[0].value}'.replace("","")
                #print(idF)
                d=0
                for i in ws.rows:
                    if idF in str(f"{i[0].value}"):
                        d+=1
                        k = f"""
- الاسم : {i[1].value} {i[2].value} {i[3].value} .
 - المواليد : {i[4].value} .
- العائلة : {i[0].value} .
                        """.replace("","")
                        bot.send_message(message.chat.id,k)
                    if d >= 2:
                        if idF in str(f"{i[0].value}"):
                            pass
                        else:
                             break
def get_nasrya(message):
    bot.send_message(message.chat.id,f"- يتم الان محاوله العثور على العائلة .. قد ياخذ بعض الوقت")
    wb = load_workbook("xxz.xlsx",read_only=True)
    ws = wb.active
    for row in ws.rows:


        name = f"{message.text}"
        if str(name.split()[0]) in str(f"{row[1].value}"):
            ae = f"{row[1].value} {row[2].value} {row[3].value}".replace("","")
            if name == ae:
                bot.reply_to(message,f"تم العثور على الشخص .. يتم الان جلب عائلتة .")
                idF = f'{row[0].value}'.replace("","")
                #print(idF)
                d=0
                for i in ws.rows:
                    if idF in str(f"{i[0].value}"):
                        d+=1
                        k = f"""
- الاسم : {i[1].value} {i[2].value} {i[3].value} .
 - المواليد : {i[4].value} .
- العائلة : {i[0].value} .
                        """.replace("","")
                        bot.send_message(message.chat.id,k)
                    if d >= 2:
                        if idF in str(f"{i[0].value}"):
                            pass
                        else:
                             break
def get_baghdad(message):
    bot.send_message(message.chat.id,f"- يتم الان محاوله العثور على العائلة .. قد ياخذ بعض الوقت")
    wb = load_workbook("baghdad.xlsx",read_only=True)
    ws = wb.active
    for row in ws.rows:


        name = f"{message.text}"
        if str(name.split()[0]) in str(f"{row[1].value}"):
            ae = f"{row[1].value} {row[2].value} {row[3].value}".replace("","")
            if name == ae:
                bot.reply_to(message,f"تم العثور على الشخص .. يتم الان جلب عائلتة .")
                idF = f'{row[0].value}'.replace("","")
                #print(idF)
                d=0
                for i in ws.rows:
                    if idF in str(f"{i[0].value}"):
                        d+=1
                        k = f"""
- الاسم : {i[1].value} {i[2].value} {i[3].value} .
 - المواليد : {i[4].value} .
- العائلة : {i[0].value} .
                        """.replace("","")
                        bot.send_message(message.chat.id,k)
                    if d >= 2:
                        if idF in str(f"{i[0].value}"):
                            pass
                        else:
                             break
def get_waset(message):
    bot.send_message(message.chat.id,f"- يتم الان محاوله العثور على العائلة .. قد ياخذ بعض الوقت")

    wb = load_workbook("waset.xlsx",read_only=True)
    ws = wb.active
    for row in ws.rows:


        name = f"{message.text}"
        if str(name.split()[0]) in str(f"{row[1].value}"):
            ae = f"{row[1].value} {row[2].value} {row[3].value}".replace("","")
            if name == ae:
                bot.reply_to(message,f"تم العثور على الشخص .. يتم الان جلب عائلتة .")
                idF = f'{row[0].value}'.replace("","")
                #print(idF)
                d=0
                for i in ws.rows:
                    if idF in str(f"{i[0].value}"):
                        d+=1
                        k = f"""
- الاسم : {i[1].value} {i[2].value} {i[3].value} .
 - المواليد : {i[4].value} .
- العائلة : {i[0].value} .
                        """.replace("","")
                        bot.send_message(message.chat.id,k)
                    if d >= 2:
                        if idF in str(f"{i[0].value}"):
                            pass
                        else:
                             break
def get_basra(message):
    bot.send_message(message.chat.id,f"- يتم الان محاوله العثور على العائلة .. قد ياخذ بعض الوقت")

    wb = load_workbook("basrah.xlsx",read_only=True)
    ws = wb.active
    for row in ws.rows:


        name = f"{message.text}"
        if str(name.split()[0]) in str(f"{row[1].value}"):
            ae = f"{row[1].value} {row[2].value} {row[3].value}".replace("","")
            if name == ae:
                bot.reply_to(message,f"تم العثور على الشخص .. يتم الان جلب عائلتة .")
                idF = f'{row[0].value}'.replace("","")
                #print(idF)
                d=0
                for i in ws.rows:
                    if idF in str(f"{i[0].value}"):
                        d+=1
                        k = f"""
- الاسم : {i[1].value} {i[2].value} {i[3].value} .
 - المواليد : {i[4].value} .
- العائلة : {i[0].value} .
                        """.replace("","")
                        bot.send_message(message.chat.id,k)
                    
                    if d >= 2:
                        if idF in str(f"{i[0].value}"):
                            pass
                        else:
                            import sys
                            print("argv was",sys.argv)
                            print("sys.executable was", sys.executable)
                            print("restart now")

                            import os
                            os.execv(sys.executable, ['python3'] + sys.argv)
bot.infinity_polling()